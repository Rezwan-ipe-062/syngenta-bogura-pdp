#!/usr/bin/env python3
"""PDP Route Optimiser - build data pack from the Bogura straight-line distance workbook.

Validates the workbook and emits browser data files:
  js/data.js        -> window.APP_DATA            (locations only; distances computed in JS)
  js/constraints.js -> window.CONSTRAINTS_REGISTER (editable Road Constraints Register seed)

Haversine distances are deliberately NOT precomputed here: the app computes the
141x141 matrix in JavaScript from the Latitude/Longitude columns. This removes any
dependency on Excel formula caching (the workbook ships with zero cached values).
"""
import json
import math
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DEFAULT_XLSX = ROOT.parent / "Data Files" / "bogura-141x141-distance-matrix-cleaned.xlsx"

E = lambda s: sys.exit(f"[build_data] FAIL: {s}")

def num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def check(cond, msg):
    if not cond:
        E(msg)

# ---------------------------------------------------------------- ask for workbook
if len(sys.argv) > 1:
    XLSX = Path(sys.argv[1]).resolve()
elif sys.stdin.isatty():
    raw = input(f"Excel workbook path [{DEFAULT_XLSX}]: ").strip()
    XLSX = Path(raw).resolve() if raw else DEFAULT_XLSX
else:
    XLSX = DEFAULT_XLSX
check(XLSX.exists(), f"workbook not found: {XLSX}")
print(f"[build_data] using workbook: {XLSX}")

# ---------------------------------------------------------------- load workbook
wb = openpyxl.load_workbook(XLSX, data_only=True)
check(set(wb.sheetnames) >= {"Read Me", "Locations", "Distance Matrix", "Pair List"},
      f"workbook missing required sheets (Read Me, Locations, Distance Matrix, Pair List); got {wb.sheetnames}")

readme = {}
for r in wb["Read Me"].iter_rows(min_row=1, max_col=2, values_only=True):
    if r[0] is not None and r[1] is not None:
        readme[str(r[0]).strip()] = str(r[1]).strip()

# ---------------------------------------------------------------- locations
rows = [r for r in wb["Locations"].iter_rows(min_row=1, max_col=6, values_only=True)]
hdr = [str(r) for r in rows[0]]
check(hdr[:6] == ["Location ID", "Location Name", "BP ID", "Address", "Latitude", "Longitude"],
      f"unexpected Locations header {hdr}")
loc_rows = [r for r in rows[1:] if r[0] is not None]
check(len(loc_rows) == 141, f"expected 141 location rows, got {len(loc_rows)}")

warehouse = None
customers = []
seen_ids = set()
seen_bp = set()
for r in loc_rows:
    lid, name, bp, addr, lat, lon = (str(r[0]).strip(), r[1], r[2], r[3], r[4], r[5])
    check(lid not in seen_ids, f"duplicate Location ID {lid}")
    seen_ids.add(lid)
    check(num(lat) and num(lon), f"{lid}: missing/invalid lat or lon ({lat!r}, {lon!r})")
    check(20.0 <= lat <= 27.0 and 88.0 <= lon <= 93.0,
          f"{lid}: coordinates outside Bangladesh bounds ({lat}, {lon})")
    if lid == "WH":
        warehouse = {"id": "WH", "name": str(name or ""), "bpId": "",
                     "address": str(addr or ""), "lat": lat, "lon": lon}
    elif re.fullmatch(r"C\d{3}", lid):
        bp = "" if bp is None else str(bp).strip()
        check(bp not in seen_bp or not bp, f"{lid}: duplicate BP ID {bp}")
        if bp:
            seen_bp.add(bp)
        customers.append({"id": lid, "name": str(name or ""), "bpId": bp,
                          "address": str(addr or ""), "lat": lat, "lon": lon})
    else:
        E(f"unexpected Location ID {lid!r}")

check(warehouse is not None, "no WH row found")
check(len(customers) == 140, f"expected 140 customers, got {len(customers)}")
customers.sort(key=lambda c: c["id"])
ids = ["WH"] + [c["id"] for c in customers]
check(set(ids) == {"WH"} | {f"C{i:03d}" for i in range(1, 141)},
      "customer IDs not the complete continuous C001..C140 set")

# ---------------------------------------------------------------- enrichment (Customer Master, display-only — optional sheet)
enr_hdr = None
enr_rows = []
if "Customer Master" in wb.sheetnames:
    for row in wb["Customer Master"].iter_rows(values_only=True):
        if row[0] is None:
            continue
        if enr_hdr is None:
            enr_hdr = [str(x).strip() if x is not None else "" for x in row]
        else:
            enr_rows.append(row)
else:
    print("[build_data] Customer Master sheet not found - region/territory enrichment skipped")

def enr_col(name):
    return enr_hdr.index(name) if name in enr_hdr else -1

if enr_hdr:
    idx_lid = enr_col("Location ID")
    idx_bp = enr_col("BP ID")
    idx_zone = enr_col("Zone Name")
    idx_unit = enr_col("Unit Name")
    idx_grp = enr_col("Sales Group Code")
    missing_cols = [n for n, v in [("Location ID", idx_lid), ("BP ID", idx_bp),
                                    ("Zone Name", idx_zone), ("Unit Name", idx_unit),
                                    ("Sales Group Code", idx_grp)] if v < 0]
    if missing_cols:
        print(f"[build_data] Customer Master missing columns {missing_cols} - region/territory enrichment skipped")
        enr_hdr = None  # disable enrichment
else:
    idx_lid = idx_bp = idx_zone = idx_unit = idx_grp = -1

enr_by_bp, enr_by_lid = {}, {}
if enr_hdr:
    for r in enr_rows:
        bp = "" if r[idx_bp] is None else str(r[idx_bp]).strip()
        lid = "" if r[idx_lid] is None else str(r[idx_lid]).strip()
        zone = "" if r[idx_zone] is None else str(r[idx_zone]).strip()
        unit = "" if r[idx_unit] is None else str(r[idx_unit]).strip()
        grp = "" if r[idx_grp] is None else str(r[idx_grp]).strip()
        rec = {"region": zone, "territory": unit, "salesGroup": grp}
        if lid:
            enr_by_lid[lid] = rec
        if bp:
            enr_by_bp[bp] = rec

    matched = 0
    for c in customers:
        rec = enr_by_lid.get(c["id"]) or enr_by_bp.get(c["bpId"], {})
        if rec.get("region"):
            matched += 1
        c["region"] = rec.get("region", "")
        c["territory"] = rec.get("territory", "")
        c["salesGroup"] = rec.get("salesGroup", "")
    print(f"[build_data] enrichment: {matched}/{len(customers)} customers matched on Location ID / BP ID")
else:
    for c in customers:
        c["region"] = ""
        c["territory"] = ""
        c["salesGroup"] = ""

# ---------------------------------------------------------------- info pack (Customer Master from sibling PDP project, display-only)
# Joins on BP ID to pull 12-month sales/volume, activity, recency and a monthly
# volume sparkline for the map. Each customer gains a compact `info` object.
INFO_SRC = ROOT.parent.parent / "PDP Route Optimization" / "pdp_app" / "output" / "master_data.json"

def _mt(mon):
    if not isinstance(mon, dict):
        return 0.0
    x = mon.get("qty_sku_mt")
    if x is None:
        x = mon.get("qty")
    return float(x or 0.0)

info_matched = 0
if INFO_SRC.exists():
    md = json.loads(INFO_SRC.read_text(encoding="utf-8"))
    info_by_bp = {}
    for m in md.get("customers", []):
        bp = "" if m.get("BP_ID") is None else str(m["BP_ID"]).strip()
        if not bp or bp in info_by_bp:
            continue
        months = m.get("Monthly") or {}
        mkeys = sorted(months.keys())[-12:]
        info_by_bp[bp] = {
            "n": str(m.get("Customer_Name") or "").strip(),
            "mk": str(m.get("Market") or "").strip(),
            "d": str(m.get("District_Parsed") or "").strip(),
            "sl": round(m.get("L12M_Sales_BDT") or 0),
            "v": round(m.get("L12M_Vol") or 0, 1),
            "a": int(m.get("Active_Months") or 0),
            "r": int(m.get("Recency_Months") or 0),
            "m": 1 if m.get("Meets_Either_Threshold") else 0,
            "sf": mkeys[0] if mkeys else "",
            "sp": [round(_mt(months.get(k)), 2) for k in mkeys],
        }
    for c in customers:
        rec = info_by_bp.get(c["bpId"])
        if rec:
            c["info"] = rec
            info_matched += 1
    print(f"[build_data] info pack: {info_matched}/{len(customers)} customers matched on BP ID from `{INFO_SRC.name}`")
else:
    print(f"[build_data] info pack: `{INFO_SRC.name}` not found - map info mode disabled")

# ---------------------------------------------------------------- matrix/pairs sanity (informative)
pair_ws = wb["Pair List"]
pairs = [r for r in pair_ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
pair_keys = set()
for r in pairs:
    a, b = str(r[0]).strip(), str(r[2]).strip()
    check(a in ids and b in ids, f"pair list references unknown ID {a}/{b}")
    if a == b:
        continue
    pair_keys.add(tuple(sorted((a, b))))
expect_pairs = {tuple(sorted((ids[i], ids[j])))
                for i in range(len(ids)) for j in range(i + 1, len(ids))}
missing_pairs = expect_pairs - pair_keys
check(not missing_pairs, f"pair list missing {len(missing_pairs)} pairs (e.g. {sorted(missing_pairs)[:3]})")
extra_pairs = pair_keys - expect_pairs
print(f"[build_data] Pair List: {len(pairs)} rows, {len(pair_keys)} unique unordered pairs, "
      f"{len(extra_pairs)} extra (ignored - app computes its own matrix)")

# ---------------------------------------------------------------- geography summary
R = 6371.0088
def hav(a, b):
    p1, p2 = math.radians(a[0]), math.radians(b[0])
    dp = math.radians(b[0] - a[0]); dl = math.radians(b[1] - a[1])
    x = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(x))

wh_km = sorted(hav((warehouse["lat"], warehouse["lon"]), (c["lat"], c["lon"])) for c in customers)
geo = {
    "minKm": round(wh_km[0], 1), "medianKm": round(wh_km[70], 1), "maxKm": round(wh_km[-1], 1),
    "beyond50km": sum(1 for d in wh_km if d > 50),
    "beyond100km": sum(1 for d in wh_km if d > 100),
    "beyond150km": sum(1 for d in wh_km if d > 150),
}

# ---------------------------------------------------------------- emit files
def js_str(o):
    return json.dumps(o, ensure_ascii=False, separators=(",", ":"))

meta = {
    "source": XLSX.name,
    "generated": "2026-08-30",
    "totalCustomers": 140,
    "routeName": "Bogura PDP",
    "geography": geo,
}
readme_txt = {
    "Purpose": readme.get("Purpose", ""),
    "Coverage": readme.get("Coverage", ""),
    "HowToUse": readme.get("How to use", ""),
    "CriticalLimitation": readme.get("Critical limitation", ""),
    "Formula": readme.get("Formula", ""),
}

data_js = (
    "/** GENERATED by build_data.py - do not edit by hand. Re-run: python build_data.py */\n"
    f"window.APP_DATA = {{\n"
    f"  meta: {js_str(meta)},\n"
    f"  readme: {js_str(readme_txt)},\n"
    f"  warehouse: {js_str(warehouse)},\n"
    f"  customers: {js_str(customers)},\n"
    f"}};\n"
)
(ROOT / "js" / "data.js").write_text(data_js, encoding="utf-8")
print(f"[build_data] wrote js/data.js ({len(data_js)} bytes)")

constraints_js = (
    "/** Road Constraints Register - editable Road/data input. Do not put <script> data in comments.\n"
    " * Every row uses Location IDs: WH or C001..C140 (From ID / To ID are a location pair).\n"
    " * Constraint Type: River / no direct crossing | Ferry required | Weak bridge |\n"
    " *   4-wheeler restricted | Seasonal / monsoon access risk | Market-time restriction |\n"
    " *   Road under repair | Security / local restriction | Other\n"
    " * Status: Blocked | Uncertain | Validated | Not reviewed\n"
    " * Rules: Blocked pair = NEVER consecutive stops. Uncertain pair = allowed only with visible warning.\n"
    " * Replace this seed with your own register via the Import button in the app (CSV/Excel/JSON).\n"
    " * This seed intentionally starts empty - constraints come from the field register.\n"
    " */\n"
    "window.CONSTRAINTS_REGISTER = [];\n"
)
(ROOT / "js" / "constraints.js").write_text(constraints_js, encoding="utf-8")
print(f"[build_data] wrote js/constraints.js ({len(constraints_js)} bytes)")

print(f"[build_data] OK. {len(customers)} customers, warehouse {warehouse['name']} "
      f"({warehouse['lat']:.4f},{warehouse['lon']:.4f}).")